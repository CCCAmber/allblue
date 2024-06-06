import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_list.cell(row=1,column=5,value="Profit")
print(product_list.max_column)

for product_row in range(2,product_list.max_row+1):
    supplier_name = product_list.cell(product_row,4).value
    product_price = product_list.cell(product_row,3).value
    product_inventory = product_list.cell(product_row,2).value
    product_profit = round(product_price*product_inventory)
    product_list.cell(row=product_row,column=5,value=product_profit)

print(product_profit)

inv_file.save("inventory_with_profit.xlsx")

print("HelloWorld")